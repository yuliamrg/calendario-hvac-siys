from __future__ import annotations

import argparse
import hashlib
import json
import re
import struct
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import Page, expect, sync_playwright


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def get_state(page: Page) -> dict:
    return page.evaluate(
        """
        async () => new Promise((resolve, reject) => {
          const databaseName = location.pathname.includes("/beta/")
            ? "calendario-hvac-siys-beta"
            : "calendario-hvac-siys";
          const request = indexedDB.open(databaseName, 1);
          request.onerror = () => reject(request.error);
          request.onsuccess = () => {
            const db = request.result;
            const tx = db.transaction("documents", "readonly");
            const get = tx.objectStore("documents").get("current");
            get.onerror = () => reject(get.error);
            get.onsuccess = () => resolve(get.result?.document ?? null);
          };
        })
        """
    )


def wait_saved(page: Page) -> None:
    expect(page.locator("#saveIndicatorText")).to_have_text("Guardado", timeout=15_000)


def click_menu_action(page: Page, button_id: str) -> None:
    menu = page.locator(f".action-menu:has(#{button_id})")
    if menu.get_attribute("open") is None:
        menu.locator("summary").click()
    page.locator(f"#{button_id}").click()


def launch_and_check(
    playwright,
    channel: str,
    html_path: Path,
    base_path: Path,
    artifact_dir: Path,
    run_full: bool,
) -> dict:
    browser = playwright.chromium.launch(channel=channel, headless=True)
    context = browser.new_context(
        accept_downloads=True,
        locale="es-CO",
        timezone_id="America/Bogota",
        viewport={"width": 1600, "height": 1000},
    )
    context.add_init_script(
        """
        (() => {
          const NativeDate = Date;
          const fixedNow = NativeDate.parse("2026-07-01T12:00:00Z");
          class FixedDate extends NativeDate {
            constructor(...args) {
              super(...(args.length ? args : [fixedNow]));
            }
            static now() {
              return fixedNow;
            }
          }
          FixedDate.parse = NativeDate.parse;
          FixedDate.UTC = NativeDate.UTC;
          window.Date = FixedDate;
        })();
        """
    )
    page = context.new_page()
    page_errors: list[str] = []
    console_errors: list[str] = []
    network_requests: list[str] = []
    page.on("pageerror", lambda error: page_errors.append(str(error)))
    page.on(
        "console",
        lambda message: console_errors.append(message.text)
        if message.type == "error"
        else None,
    )
    page.on(
        "request",
        lambda request: network_requests.append(request.url)
        if request.url.startswith(("http://", "https://"))
        else None,
    )

    page.goto(html_path.as_uri(), wait_until="load")
    page.wait_for_selector('body[data-ready="true"]', timeout=20_000)
    expect(page).to_have_title("SIYS Sync")
    expect(page.locator("html")).to_have_attribute("data-theme", "light")
    expect(page.get_by_test_id("month-title")).to_have_text("Julio de 2026")
    weekdays = page.locator("#weekdayRow > div").all_inner_texts()
    assert weekdays == [
        "Lunes",
        "Martes",
        "Miércoles",
        "Jueves",
        "Viernes",
        "Sábado",
        "Domingo",
    ], weekdays
    assert page.locator(".calendar-grid .day-cell").count() == 42
    assert page.locator(".calendar-grid .day-number").count() == 42
    assert page.locator('.calendar-grid .day-number[tabindex="0"]').count() == 1
    page.locator('[data-date="2026-07-30"] .day-number').focus()
    page.keyboard.press("ArrowLeft")
    assert page.evaluate(
        "document.activeElement.closest('.day-cell')?.dataset.date"
    ) == "2026-07-29"
    for dialog_id in (
        "activityDialog",
        "catalogDialog",
        "importDialog",
        "filterDialog",
        "holidayDialog",
        "bulkMoveDialog",
        "bulkStatusDialog",
        "bulkEditDialog",
        "restoreDialog",
        "programmingImportDialog",
        "calendarSettingsDialog",
        "dropActionDialog",
        "resetDataDialog",
        "mergeJsonDialog",
        "helpDialog",
    ):
        labelled_by = page.locator(f"#{dialog_id}").get_attribute("aria-labelledby")
        assert labelled_by and page.locator(f"#{labelled_by}").count() == 1
    expect(page.locator("#detailDrawer")).not_to_be_visible()
    expect(page.locator('[data-date="2026-07-13"]')).to_have_class(
        re.compile(r"\bholiday\b")
    )
    expect(page.locator('[data-date="2026-07-12"]')).to_have_class(
        re.compile(r"\bsunday\b")
    )

    if not run_full:
        page.screenshot(path=str(artifact_dir / f"{channel}-smoke.png"), full_page=True)
        assert not page_errors, page_errors
        assert not console_errors, console_errors
        assert not network_requests, network_requests
        context.close()
        browser.close()
        return {"channel": channel, "status": "ok", "mode": "smoke"}

    second_page = context.new_page()
    second_page.goto(html_path.as_uri(), wait_until="load")
    second_page.wait_for_selector('body[data-ready="true"]', timeout=20_000)
    expect(second_page.locator("#accessBanner")).to_be_visible()
    expect(second_page.locator("#newActivityButton")).to_be_disabled()
    second_page.on("dialog", lambda dialog: dialog.accept())
    second_page.locator("#takeControlButton").click()
    expect(second_page.locator("#accessBanner")).not_to_be_visible()
    expect(page.locator("#accessBanner")).to_be_visible()
    page.on("dialog", lambda dialog: dialog.accept())
    page.locator("#takeControlButton").click()
    expect(page.locator("#accessBanner")).not_to_be_visible()
    expect(second_page.locator("#accessBanner")).to_be_visible()
    second_page.close()
    expect(page.locator("#storageBanner")).to_be_visible()

    manage_menu = page.locator(".action-menu", has_text="Gestionar")
    share_menu = page.locator(".action-menu", has_text="Compartir")
    manage_menu.locator("summary").click()
    expect(manage_menu).to_have_attribute("open", "")
    share_menu.locator("summary").click()
    expect(manage_menu).not_to_have_attribute("open", "")
    expect(share_menu).to_have_attribute("open", "")
    page.keyboard.press("Escape")
    expect(share_menu).not_to_have_attribute("open", "")

    click_menu_action(page, "calendarSettingsButton")
    page.fill("#calendarName", "Cronograma automatizado")
    page.fill("#calendarCoordinator", "Coordinación QA")
    page.locator("#calendarSettingsForm button[type=submit]").click()
    wait_saved(page)
    identified_state = get_state(page)
    assert identified_state["schemaVersion"] == 4
    assert identified_state["calendarMeta"]["name"] == "Cronograma automatizado"
    assert identified_state["calendarMeta"]["coordinator"] == "Coordinación QA"
    assert identified_state["calendarMeta"]["revision"] == 1

    click_menu_action(page, "themeButton")
    expect(page.locator("html")).to_have_attribute("data-theme", "light")
    click_menu_action(page, "themeButton")
    expect(page.locator("html")).to_have_attribute("data-theme", "dark")
    assert get_state(page)["calendarMeta"]["revision"] == 1
    theme_page = context.new_page()
    theme_page.goto(html_path.as_uri(), wait_until="load")
    theme_page.wait_for_selector('body[data-ready="true"]', timeout=20_000)
    expect(theme_page.locator("html")).to_have_attribute("data-theme", "dark")
    assert get_state(theme_page)["calendarMeta"]["revision"] == 1
    theme_page.close()
    page.screenshot(path=str(artifact_dir / f"{channel}-dark-theme.png"), full_page=True)
    click_menu_action(page, "themeButton")
    expect(page.locator("html")).to_have_attribute("data-theme", "light")
    page.locator(".action-menu", has_text="Gestionar").locator("summary").click()
    contrast_failures = page.evaluate(
        """
        () => {
          const parse = value => (value.match(/[\\d.]+/g) || []).map(Number);
          const luminance = rgb => {
            const values = rgb.slice(0, 3).map(value => {
              const channel = value / 255;
              return channel <= 0.04045 ? channel / 12.92 : ((channel + 0.055) / 1.055) ** 2.4;
            });
            return 0.2126 * values[0] + 0.7152 * values[1] + 0.0722 * values[2];
          };
          const opaqueBackground = element => {
            let node = element;
            while (node) {
              const color = getComputedStyle(node).backgroundColor;
              const parsed = parse(color);
              if ((parsed[3] ?? 1) >= 0.95) return parsed;
              node = node.parentElement;
            }
            return [16, 23, 19, 1];
          };
          return [...document.querySelectorAll(".button, .menu-action, .segment, .day-number, input")]
            .filter(element => element.getClientRects().length && !element.disabled)
            .map(element => {
              const foreground = parse(getComputedStyle(element).color);
              const background = opaqueBackground(element);
              const light = Math.max(luminance(foreground), luminance(background));
              const dark = Math.min(luminance(foreground), luminance(background));
              return { label: element.textContent?.trim() || element.placeholder || element.id, ratio: (light + 0.05) / (dark + 0.05) };
            })
            .filter(item => item.ratio < 3);
        }
        """
    )
    assert not contrast_failures, contrast_failures
    page.keyboard.press("Escape")

    system_context = browser.new_context(locale="es-CO", color_scheme="dark")
    system_context.add_init_script(
        """
        localStorage.setItem(
          `siys-sync-ui:${location.pathname.includes("/beta/") ? "beta" : "local"}`,
          JSON.stringify({theme: "system"})
        );
        """
    )
    system_page = system_context.new_page()
    system_page.goto(html_path.as_uri(), wait_until="load")
    system_page.wait_for_selector('body[data-ready="true"]', timeout=20_000)
    expect(system_page.locator("html")).to_have_attribute("data-theme", "dark")
    system_context.close()

    before_hash = file_hash(base_path)
    page.set_input_files("#baseFileInput", str(base_path))
    expect(page.locator("#importDialog")).to_be_visible(timeout=20_000)
    expect(page.locator("#importFileSummary")).to_contain_text(base_path.name)
    page.get_by_test_id("apply-import").click()
    expect(page.locator("#importDialog")).not_to_be_visible()
    wait_saved(page)
    after_hash = file_hash(base_path)
    assert before_hash == after_hash

    state = get_state(page)
    assert len(state["catalog"]["clients"]) == 3
    assert len(state["catalog"]["sites"]) == 18
    assert len(state["catalog"]["responsibles"]) >= 30
    serialized = json.dumps(state, ensure_ascii=False)
    for forbidden_key in [
        "cedula_nit",
        '"Correo"',
        '"contacto"',
        '"Contactos"',
        "created_by",
        '"photos"',
        '"files"',
    ]:
        assert forbidden_key not in serialized

    client = next(
        item for item in state["catalog"]["clients"] if item["name"] == "Coopidrogas"
    )
    site = next(
        item
        for item in state["catalog"]["sites"]
        if item["clientId"] == client["id"]
        and item.get("active", True)
        and item.get("city")
        and item["city"] != "No aplica"
    )
    payroll = next(
        item
        for item in state["catalog"]["responsibles"]
        if item["responsibleType"] == "payroll" and item.get("active", True)
    )
    contractor = next(
        item
        for item in state["catalog"]["responsibles"]
        if item["responsibleType"] == "contractor" and item.get("active", True)
    )

    site_card = page.locator(
        f'[data-drag-type="site"][data-site-id="{site["id"]}"]'
    )
    expect(site_card).to_be_visible()
    page.evaluate(
        """
        ({ siteId, date }) => {
          const source = document.querySelector(`[data-drag-type="site"][data-site-id="${siteId}"]`);
          const target = document.querySelector(`[data-date="${date}"]`);
          const dataTransfer = new DataTransfer();
          source.dispatchEvent(new DragEvent("dragstart", { bubbles: true, dataTransfer }));
          target.dispatchEvent(new DragEvent("drop", { bubbles: true, dataTransfer }));
        }
        """,
        {"siteId": site["id"], "date": "2026-07-30"},
    )
    expect(page.get_by_test_id("activity-dialog")).to_be_visible()
    actual_client = page.locator("#activityClient").input_value()
    actual_site = page.locator("#activitySite").input_value()
    assert actual_client == client["id"], {
        "expectedClient": client["id"],
        "actualClient": actual_client,
        "expectedSite": site["id"],
        "actualSite": actual_site,
    }
    assert actual_site == site["id"]
    page.select_option("#activityStatus", "confirmed")
    page.locator(
        f'#responsiblePicker input[value="{payroll["id"]}"]'
    ).check()
    page.locator(
        f'#responsiblePicker input[value="{contractor["id"]}"]'
    ).check()
    page.fill("#activityObservations", "Prueba integral local")
    page.get_by_test_id("save-activity").click()
    expect(page.get_by_test_id("activity-dialog")).not_to_be_visible()
    wait_saved(page)
    state = get_state(page)
    assert len(state["activities"]) == 1
    activity_id = state["activities"][0]["id"]
    card = page.locator(f'[data-activity-id="{activity_id}"]')
    expect(card).to_have_class(re.compile(r"\bmixed\b"))
    expect(card.locator(".service-code")).to_have_text("MP")
    expect(card).to_have_attribute("data-service-code", "MP")
    expect(card).to_have_attribute("aria-label", re.compile(r"tipo de servicio: Mantenimiento preventivo"))
    if "open" in (page.locator("#detailDrawer").get_attribute("class") or ""):
        page.locator("#closeDrawerButton").click()
        page.wait_for_timeout(300)

    original_revision = get_state(page)["calendarMeta"]["revision"]
    original_history_count = len(get_state(page)["activities"][0]["history"])
    page.evaluate(
        """
        ({ activityId, date }) => {
          const source = document.querySelector(`[data-activity-id="${activityId}"]`);
          const target = document.querySelector(`[data-date="${date}"]`);
          const dataTransfer = new DataTransfer();
          source.dispatchEvent(new DragEvent("dragstart", { bubbles: true, dataTransfer }));
          target.dispatchEvent(new DragEvent("drop", { bubbles: true, dataTransfer }));
          source.dispatchEvent(new DragEvent("dragend", { bubbles: true, dataTransfer }));
        }
        """,
        {"activityId": activity_id, "date": "2026-07-30"},
    )
    expect(page.locator("#dropActionDialog")).not_to_be_visible()
    same_day_state = get_state(page)
    assert same_day_state["calendarMeta"]["revision"] == original_revision
    assert len(same_day_state["activities"][0]["history"]) == original_history_count

    target_locator = page.locator('[data-date="2026-07-28"]')
    page.evaluate(
        """
        ({ activityId, date }) => {
          const source = document.querySelector(`[data-activity-id="${activityId}"]`);
          const target = document.querySelector(`[data-date="${date}"]`);
          const dataTransfer = new DataTransfer();
          source.dispatchEvent(new DragEvent("dragstart", { bubbles: true, dataTransfer }));
          target.dispatchEvent(new DragEvent("drop", { bubbles: true, dataTransfer }));
          source.dispatchEvent(new DragEvent("dragend", { bubbles: true, dataTransfer }));
        }
        """,
        {"activityId": activity_id, "date": "2026-07-28"},
    )
    expect(page.locator("#dropActionDialog")).to_be_visible()
    page.locator("#dropMoveButton").click()
    wait_saved(page)
    state = get_state(page)
    activity = next(item for item in state["activities"] if item["id"] == activity_id)
    assert activity["date"] == "2026-07-28", {
        "activity": activity,
        "toasts": page.locator("#toastRegion").inner_text(),
        "targetClass": target_locator.get_attribute("class"),
    }
    assert any(item["action"] == "rescheduled" for item in activity["history"])

    page.locator(f'[data-activity-id="{activity_id}"]').click()
    expect(page.locator("#detailDrawer")).to_have_class(re.compile(r"\bopen\b"))
    expect(page.locator("#detailDrawer")).to_be_visible()
    drawer_box = page.locator("#detailDrawer").bounding_box()
    assert drawer_box is not None
    assert abs((drawer_box["x"] + drawer_box["width"] / 2) - 800) < 3
    expect(page.locator("#drawerBody")).to_contain_text("Prueba integral local")
    expect(page.locator("#drawerStatusSelect")).to_have_attribute(
        "aria-label", "Nuevo estado de la actividad"
    )
    page.select_option("#drawerStatusSelect", "completed")
    page.get_by_role("button", name="Aplicar estado").click()
    wait_saved(page)
    expect(page.locator(f'[data-activity-id="{activity_id}"]')).to_have_class(
        re.compile(r"\bcompleted\b")
    )
    opacity = page.locator(
        f'[data-activity-id="{activity_id}"]'
    ).evaluate("element => getComputedStyle(element).opacity")
    assert float(opacity) < 0.7
    page.locator("#closeDrawerButton").click()
    expect(page.locator("#detailDrawer")).not_to_be_visible()

    page.get_by_test_id("new-activity").click()
    page.fill("#activityDate", "2026-07-09")
    page.fill("#activityEndDate", "2026-07-14")
    page.select_option("#activityServiceType", "administrative")
    page.fill("#activityObservations", "Actividad administrativa multidía")
    expect(page.locator("#rangePreview")).to_contain_text("4 tarjetas")
    expect(page.locator("#rangePreview")).to_contain_text("2 fechas")
    page.get_by_test_id("save-activity").click()
    expect(page.get_by_test_id("activity-dialog")).not_to_be_visible()
    wait_saved(page)

    state = get_state(page)
    assert len(state["activities"]) == 5
    series_id = next(
        item["seriesId"]
        for item in state["activities"]
        if item.get("seriesId") is not None
    )
    series_items = sorted(
        [item for item in state["activities"] if item.get("seriesId") == series_id],
        key=lambda item: item["date"],
    )
    assert [item["date"] for item in series_items] == [
        "2026-07-09",
        "2026-07-10",
        "2026-07-11",
        "2026-07-14",
    ]
    assert len({item["id"] for item in series_items}) == 4

    if page.locator("#detailDrawer").is_visible():
        page.locator("#closeDrawerButton").click()
    for item in series_items[:2]:
        page.locator(
            f'[data-activity-id="{item["id"]}"] .activity-select'
        ).check()
    expect(page.locator("#selectionBar")).to_be_visible()
    page.locator("#bulkMoveButton").click()
    page.fill("#bulkMoveDate", "2026-07-21")
    page.locator("#bulkMoveForm button[type=submit]").click()
    expect(page.locator("#bulkMoveDialog")).not_to_be_visible()
    wait_saved(page)
    state = get_state(page)
    moved_series = sorted(
        [item for item in state["activities"] if item.get("seriesId") == series_id],
        key=lambda item: item["date"],
    )
    assert {item["date"] for item in moved_series}.issuperset(
        {"2026-07-21", "2026-07-22"}
    )

    if "open" in (page.locator("#detailDrawer").get_attribute("class") or ""):
        page.locator("#closeDrawerButton").click()
    editable_items = moved_series[:2]
    for item in editable_items:
        page.locator(f'[data-activity-id="{item["id"]}"] .activity-select').check()
    page.locator("#bulkEditButton").click()
    page.select_option("#bulkEditField", "observations")
    page.select_option("#bulkEditMode", "append")
    page.fill("#bulkEditTextarea", "Nota múltiple")
    page.locator("#bulkEditForm button[type=submit]").click()
    wait_saved(page)
    state = get_state(page)
    for item in editable_items:
        changed = next(activity for activity in state["activities"] if activity["id"] == item["id"])
        assert "Nota múltiple" in changed["observations"]
        assert changed["history"][-1]["action"] == "bulk_edited"

    for item in editable_items:
        page.locator(f'[data-activity-id="{item["id"]}"] .activity-select').check()
    activity_count_before_delete = len(get_state(page)["activities"])
    page.locator("#bulkDeleteButton").click()
    wait_saved(page)
    assert len(get_state(page)["activities"]) == activity_count_before_delete - 2
    page.get_by_role("button", name="Deshacer").last.click()
    wait_saved(page)
    assert len(get_state(page)["activities"]) == activity_count_before_delete

    anchor = next(item for item in moved_series if item["date"] == "2026-07-21")
    page.locator(f'[data-activity-id="{anchor["id"]}"]').click()
    page.select_option("#drawerStatusSelect", "completed")
    page.locator(
        'input[name="drawerStatusScope"][value="series"]'
    ).check()
    page.get_by_role("button", name="Aplicar estado").click()
    wait_saved(page)
    state = get_state(page)
    assert all(
        item["status"] == "completed"
        for item in state["activities"]
        if item.get("seriesId") == series_id
    )
    page.locator("#closeDrawerButton").click()

    click_menu_action(page, "holidayButton")
    page.fill("#overrideDate", "2026-07-29")
    page.select_option("#overrideType", "manual-closure")
    page.fill("#overrideName", "Cierre de prueba")
    page.fill("#overrideReason", "Validación automatizada")
    page.locator("#holidayForm button[type=submit]").click()
    expect(page.locator("#overrideList")).to_contain_text("Cierre de prueba")
    page.locator('[data-close-dialog="holidayDialog"]').last.click()
    expect(page.locator('[data-date="2026-07-29"]')).to_have_class(
        re.compile(r"\bholiday\b")
    )
    wait_saved(page)

    state_before_reload = get_state(page)
    page.reload(wait_until="load")
    page.wait_for_selector('body[data-ready="true"]', timeout=20_000)
    state_after_reload = get_state(page)
    assert len(state_after_reload["activities"]) == len(
        state_before_reload["activities"]
    )
    assert len(state_after_reload["holidayOverrides"]) == 1

    with page.expect_download() as download_info:
            click_menu_action(page, "backupButton")
    backup_download = download_info.value
    assert re.match(
        r"\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}_respaldo-cronograma_[a-z0-9-]+\.json",
        backup_download.suggested_filename,
    )
    backup_path = artifact_dir / "respaldo-prueba.json"
    backup_download.save_as(str(backup_path))
    backup = json.loads(backup_path.read_text(encoding="utf-8"))
    assert backup["format"] == "calendario-hvac-siys-backup"
    assert backup["revision"] == state_after_reload["calendarMeta"]["revision"]
    backup_document = backup["document"]
    assert len(backup_document["activities"]) == len(state_after_reload["activities"])

    with page.expect_download() as download_info:
            click_menu_action(page, "exportCsvButton")
    csv_download = download_info.value
    assert re.match(
        r"\d{4}-\d{2}_programacion_[a-z0-9-]+\.csv",
        csv_download.suggested_filename,
    )
    csv_path = artifact_dir / "programacion-prueba.csv"
    csv_download.save_as(str(csv_path))
    csv_text = csv_path.read_text(encoding="utf-8-sig")
    assert "Responsables nómina" in csv_text
    assert "Responsables contratistas" in csv_text
    assert "Prueba integral local" in csv_text
    assert "cedula_nit" not in csv_text

    page.get_by_test_id("new-activity").click()
    page.fill("#activityDate", "2026-07-28")
    page.select_option("#activityServiceType", "administrative")
    page.fill("#activityObservations", "Temporal para probar restauración")
    page.get_by_test_id("save-activity").click()
    wait_saved(page)
    assert len(get_state(page)["activities"]) == len(backup_document["activities"]) + 1

    page.set_input_files("#restoreFileInput", str(backup_path))
    expect(page.locator("#restoreDialog")).to_be_visible()
    expect(page.locator("#restoreWarning")).to_contain_text("más antiguo")
    page.locator("#restoreForm button[type=submit]").click()
    expect(page.locator("#restoreDialog")).not_to_be_visible()
    wait_saved(page)
    assert len(get_state(page)["activities"]) == len(backup_document["activities"])

    with page.expect_download() as download_info:
            click_menu_action(page, "programmingTemplateButton")
    assert download_info.value.suggested_filename == "plantilla_programacion_SIYS-Sync.xlsx"
    template_path = artifact_dir / "plantilla-programacion-hvac.xlsx"
    download_info.value.save_as(str(template_path))
    workbook = load_workbook(template_path)
    assert workbook.sheetnames == ["Programacion", "Catalogos", "Instrucciones"]
    programming_sheet = workbook["Programacion"]
    assert [cell.value for cell in programming_sheet[1]] == [
        "FechaInicio", "FechaFin", "Bandeja", "Cliente", "Sede", "Ciudad", "Responsables",
        "TipoServicio", "Estado", "Observaciones", "IncluirNoLaborables",
    ]
    programming_sheet.delete_rows(2, programming_sheet.max_row)
    programming_sheet.append([
        "2026-07-31", "2026-07-31", "Calendario", client["name"], site["name"], site["city"],
        payroll["name"], "Mantenimiento preventivo", "Programada",
        "Importada desde plantilla", "No",
    ])
    import_path = artifact_dir / "programacion-importar.xlsx"
    workbook.save(import_path)
    count_before_programming_import = len(get_state(page)["activities"])
    page.set_input_files("#programmingFileInput", str(import_path))
    expect(page.locator("#programmingImportDialog")).to_be_visible()
    expect(page.locator("#programmingImportStats")).to_contain_text("Filas válidas")
    page.locator("#programmingImportForm button[type=submit]").click()
    wait_saved(page)
    imported_state = get_state(page)
    assert len(imported_state["activities"]) == count_before_programming_import + 1
    assert any(
        activity["observations"] == "Importada desde plantilla"
        for activity in imported_state["activities"]
    )

    page.locator("#filterButton").click()
    city_filters = page.locator('input[name="filter-cities"]:not(:disabled)')
    assert city_filters.count() >= 1
    city_filters.nth(0).check()
    page.locator("#filterForm button[type=submit]").click()
    expect(page.locator("#filterCount")).to_have_text("1")
    with page.expect_download() as download_info:
            click_menu_action(page, "exportImageButton")
    png_path = artifact_dir / "cronograma-filtrado.png"
    download_info.value.save_as(str(png_path))
    wait_saved(page)
    png_data = png_path.read_bytes()
    assert png_data[:8] == b"\x89PNG\r\n\x1a\n"
    width, height = struct.unpack(">II", png_data[16:24])
    assert width >= 2000
    assert height >= 1000

    page.evaluate(
        """
        async () => new Promise((resolve, reject) => {
          const databaseName = location.pathname.includes("/beta/")
            ? "calendario-hvac-siys-beta"
            : "calendario-hvac-siys";
          const request = indexedDB.open(databaseName, 1);
          request.onerror = () => reject(request.error);
          request.onsuccess = () => {
            const db = request.result;
            const tx = db.transaction("documents", "readwrite");
            const store = tx.objectStore("documents");
            const getCurrent = store.get("current");
            getCurrent.onsuccess = () => {
              const recoveryDocument = structuredClone(getCurrent.result.document);
              recoveryDocument.calendarMeta.name = "Cronograma recuperado QA";
              store.put({
                key: "recovery",
                savedAt: new Date().toISOString(),
                document: recoveryDocument
              });
              store.put({
              key: "current",
              savedAt: new Date().toISOString(),
              document: { schemaVersion: 999, appVersion: "corrupta" }
              });
            };
            tx.oncomplete = resolve;
            tx.onerror = () => reject(tx.error);
          };
        })
        """
    )
    page.reload(wait_until="load")
    wait_ready_state = page.wait_for_selector('body[data-ready="true"]', timeout=20_000)
    assert wait_ready_state
    recovered_state = get_state(page)
    assert recovered_state["schemaVersion"] == 4
    assert recovered_state["appVersion"] != "corrupta"
    assert recovered_state["calendarMeta"]["name"] == "Cronograma recuperado QA"

    page.screenshot(path=str(artifact_dir / f"{channel}-full.png"), full_page=True)
    assert not page_errors, page_errors
    assert not console_errors, console_errors
    assert not network_requests, network_requests

    context.close()
    browser.close()
    return {
        "channel": channel,
        "status": "ok",
        "mode": "full",
        "catalog": {
            "clients": len(state["catalog"]["clients"]),
            "sites": len(state["catalog"]["sites"]),
            "responsibles": len(state["catalog"]["responsibles"]),
        },
        "activities": len(backup_document["activities"]),
        "baseSha256": before_hash,
        "networkRequests": 0,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--html", required=True, type=Path)
    parser.add_argument("--base", required=True, type=Path)
    parser.add_argument("--artifacts", type=Path)
    args = parser.parse_args()

    html_path = args.html.resolve()
    base_path = args.base.resolve()
    artifact_dir = (
        args.artifacts.resolve()
        if args.artifacts
        else Path(tempfile.mkdtemp(prefix="calendario-hvac-qa-"))
    )
    artifact_dir.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as playwright:
        results = [
            launch_and_check(
                playwright,
                "chrome",
                html_path,
                base_path,
                artifact_dir,
                run_full=True,
            ),
            launch_and_check(
                playwright,
                "msedge",
                html_path,
                base_path,
                artifact_dir,
                run_full=False,
            ),
        ]

    print(
        json.dumps(
            {
                "status": "ok",
                "results": results,
                "artifacts": str(artifact_dir),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
